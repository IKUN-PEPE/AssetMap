import argparse
import asyncio
import csv
import logging
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import async_playwright

PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = PROJECT_DIR / "screenshots"
DEFAULT_RESULT_CSV = PROJECT_DIR / "results" / "results.csv"
DEFAULT_SUMMARY_TXT = PROJECT_DIR / "results" / "summary.txt"
RESULT_HEADERS = ["seq", "input", "final_url", "status", "error", "screenshot_path"]
RESULT_STATUS_SUCCESS = "success"
RESULT_STATUS_FAILED = "failed"
RESULT_STATUS_SKIPPED = "skipped"
LOGGER_NAME = "asset_mapping_screenshot_xlsx"


def build_runtime_paths(base_dir: Path | None = None) -> dict[str, Path]:
    root = base_dir or PROJECT_DIR
    return {
        "base_dir": root,
        "log_path": root / "run.log",
        "output_dir": root / "screenshots",
        "result_csv": root / "results" / "results.csv",
        "summary_txt": root / "results" / "summary.txt",
    }


def setup_logger(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.propagate = False
    return logger


def build_run_config(
    input_path: Path,
    base_dir: Path | None = None,
    max_items: int = 0,
    concurrency: int = 5,
    retry_count: int = 1,
    timeout_sec: int = 15,
    wait_after_load: int = 5,
    skip_existing: bool = True,
    headful: bool = False,
) -> dict[str, object]:
    paths = build_runtime_paths(base_dir)
    return {
        "input_path": input_path,
        "output_dir": paths["output_dir"],
        "result_csv": paths["result_csv"],
        "summary_txt": paths["summary_txt"],
        "log_path": paths["log_path"],
        "max_items": max_items,
        "concurrency": concurrency,
        "retry_count": retry_count,
        "timeout_sec": timeout_sec,
        "wait_after_load": wait_after_load,
        "skip_existing": skip_existing,
        "headful": headful,
    }


def select_target_value(row: dict[str, str]) -> str:
    url_value = str(row.get("URL链接", "") or "").strip()
    if url_value:
        return url_value
    return str(row.get("域名/主机", "") or "").strip()


def build_candidate_urls(raw_value: str) -> list[str]:
    value = (raw_value or "").strip()
    if not value:
        return []
    if value.startswith("http://") or value.startswith("https://"):
        return [value]
    return [f"https://{value}", f"http://{value}"]


def sanitize_filename(text: str, limit: int = 120) -> str:
    cleaned = str(text or "").strip()
    for char in '<>:"/\\|?*':
        cleaned = cleaned.replace(char, "_")
    return cleaned[:limit] or "site"


def build_output_filename(seq: str, page_title: str, final_url: str) -> str:
    safe_seq = sanitize_filename(seq, limit=20)
    safe_title = sanitize_filename(page_title or "未命名", limit=80)
    safe_url = sanitize_filename(final_url, limit=160)
    return f"{safe_seq}_{safe_title}_{safe_url}.png"


def build_result_row(seq: str, input_value: str, final_url: str, status: str, error: str, screenshot_path: Path) -> dict[str, str]:
    return {
        "seq": seq,
        "input": input_value,
        "final_url": final_url,
        "status": status,
        "error": error,
        "screenshot_path": str(screenshot_path),
    }


def classify_failure_reason(error: str) -> str:
    lowered = (error or "").lower()
    if "timeout" in lowered:
        return "timeout"
    if "name_not_resolved" in lowered or "err_name_not_resolved" in lowered or "connection" in lowered or "dns" in lowered:
        return "dns_or_network"
    if "ssl" in lowered or "cert" in lowered:
        return "ssl"
    return "other"


def build_summary(results: list[dict[str, str]]) -> dict[str, object]:
    total = len(results)
    success = sum(1 for item in results if item.get("status") == RESULT_STATUS_SUCCESS)
    failed = sum(1 for item in results if item.get("status") == RESULT_STATUS_FAILED)
    skipped = sum(1 for item in results if item.get("status") == RESULT_STATUS_SKIPPED)
    reasons = Counter(classify_failure_reason(item.get("error", "")) for item in results if item.get("status") == RESULT_STATUS_FAILED)
    return {
        "total": total,
        "success": success,
        "failed": failed,
        "skipped": skipped,
        "success_rate": f"{(success / total * 100) if total else 0:.2f}%",
        "failed_rate": f"{(failed / total * 100) if total else 0:.2f}%",
        "skipped_rate": f"{(skipped / total * 100) if total else 0:.2f}%",
        "failure_reasons": dict(reasons),
    }


def render_summary_text(summary: dict[str, object]) -> str:
    lines = [
        f"Total: {summary['total']}",
        f"Success: {summary['success']} ({summary['success_rate']})",
        f"Failed: {summary['failed']} ({summary['failed_rate']})",
        f"Skipped: {summary['skipped']} ({summary['skipped_rate']})",
        "Failure reasons:",
    ]
    failure_reasons = summary.get("failure_reasons", {})
    if failure_reasons:
        for reason, count in sorted(failure_reasons.items()):
            lines.append(f"- {reason}: {count}")
    else:
        lines.append("- none")
    return "\n".join(lines)


def write_summary_text(summary_path: Path, summary_text: str) -> None:
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(summary_text, encoding="utf-8")


def finalize_screenshot_path(current_path: Path, seq: str, fallback_title: str, actual_title: str, final_url: str) -> Path:
    chosen_title = (actual_title or "").strip() or fallback_title or "未命名"
    target_path = current_path.with_name(build_output_filename(seq=seq, page_title=chosen_title, final_url=final_url))
    if target_path == current_path:
        return current_path
    if target_path.exists():
        target_path.unlink()
    current_path.replace(target_path)
    return target_path


def load_assets_from_xlsx(workbook_path: Path, max_items: int) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    required_headers = {"序号", "域名/主机", "网页标题", "URL链接"}
    headers: list[str] | None = None

    for row in rows:
        normalized = [str(cell or "").strip() for cell in row]
        if required_headers.issubset(set(normalized)):
            headers = normalized
            break

    if headers is None:
        workbook.close()
        raise ValueError("Required headers not found in workbook")

    index = {header: pos for pos, header in enumerate(headers)}
    assets: list[dict[str, str]] = []
    for row in rows:
        if max_items and len(assets) >= max_items:
            break
        asset = {
            "seq": str(row[index["序号"]] or "").strip() if index["序号"] < len(row) else "",
            "host": str(row[index["域名/主机"]] or "").strip() if index["域名/主机"] < len(row) else "",
            "title": str(row[index["网页标题"]] or "").strip() if index["网页标题"] < len(row) else "",
            "url": str(row[index["URL链接"]] or "").strip() if index["URL链接"] < len(row) else "",
        }
        if any(asset.values()):
            assets.append(asset)
    workbook.close()
    return assets


def build_output_path(out_dir: Path, seq: str, page_title: str, final_url: str) -> Path:
    return out_dir / build_output_filename(seq=seq, page_title=page_title, final_url=final_url)


def should_skip_existing(output_path: Path) -> bool:
    return output_path.exists()


def process_asset_row(asset: dict[str, str], out_dir: Path, capture_func, skip_existing: bool) -> dict[str, str]:
    target_value = select_target_value({"URL链接": asset.get("url", ""), "域名/主机": asset.get("host", "")})
    candidates = build_candidate_urls(target_value)
    first_candidate = candidates[0] if candidates else target_value
    output_path = build_output_path(out_dir, asset.get("seq", ""), asset.get("title", ""), first_candidate)
    if skip_existing and should_skip_existing(output_path):
        return build_result_row(asset.get("seq", ""), target_value, first_candidate, RESULT_STATUS_SKIPPED, "", output_path)
    try:
        capture_result = capture_func(candidates, output_path)
        if isinstance(capture_result, dict):
            final_url = capture_result.get("final_url", first_candidate)
            actual_title = capture_result.get("actual_title", "")
        else:
            final_url = capture_result
            actual_title = ""
        final_path = finalize_screenshot_path(
            current_path=output_path,
            seq=asset.get("seq", ""),
            fallback_title=asset.get("title", ""),
            actual_title=actual_title,
            final_url=final_url,
        )
        return build_result_row(asset.get("seq", ""), target_value, final_url, RESULT_STATUS_SUCCESS, "", final_path)
    except Exception as exc:
        return build_result_row(asset.get("seq", ""), target_value, first_candidate, RESULT_STATUS_FAILED, str(exc), output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="域名资产合并汇总表（2026-04-11）.xlsx")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--result-csv", default=str(DEFAULT_RESULT_CSV))
    parser.add_argument("--max-items", type=int, default=0)
    parser.add_argument("--timeout-sec", type=int, default=15)
    parser.add_argument("--wait-after-load", type=int, default=5)
    parser.add_argument("--concurrency", type=int, default=5)
    parser.add_argument("--retry-count", type=int, default=1)
    parser.add_argument("--skip-existing", dest="skip_existing", action="store_true")
    parser.add_argument("--no-skip-existing", dest="skip_existing", action="store_false")
    parser.add_argument("--headful", action="store_true")
    parser.set_defaults(skip_existing=True)
    return parser


def write_results_csv(results: list[dict[str, str]], result_csv: Path) -> None:
    result_csv.parent.mkdir(parents=True, exist_ok=True)
    with result_csv.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RESULT_HEADERS)
        writer.writeheader()
        writer.writerows(results)


async def capture_screenshot(page, candidate_urls: list[str], output_path: Path, timeout_sec: int, wait_after_load: int) -> dict[str, str]:
    last_error: Exception | None = None
    output_path.parent.mkdir(parents=True, exist_ok=True)
    for url in candidate_urls:
        try:
            await page.goto(url, timeout=timeout_sec * 1000, wait_until="domcontentloaded")
            await asyncio.sleep(wait_after_load)
            await page.screenshot(path=str(output_path), full_page=False)
            return {
                "final_url": url,
                "actual_title": (await page.title()).strip(),
            }
        except Exception as exc:
            last_error = exc
    if last_error is not None:
        raise last_error
    raise ValueError("No candidate URLs generated")


async def run_one_asset(asset: dict[str, str], browser_context, semaphore: asyncio.Semaphore, out_dir: Path, timeout_sec: int, wait_after_load: int, retry_count: int, skip_existing: bool, logger: logging.Logger | None = None) -> dict[str, str]:
    async with semaphore:
        target_value = select_target_value({"URL链接": asset.get("url", ""), "域名/主机": asset.get("host", "")})
        candidates = build_candidate_urls(target_value)
        first_candidate = candidates[0] if candidates else target_value
        output_path = build_output_path(out_dir, asset.get("seq", ""), asset.get("title", ""), first_candidate)

        if skip_existing and should_skip_existing(output_path):
            result = build_result_row(asset.get("seq", ""), target_value, first_candidate, RESULT_STATUS_SKIPPED, "", output_path)
            if logger:
                logger.info("Skipped asset %s -> %s", asset.get("seq", ""), first_candidate)
            return result

        page = await browser_context.new_page()
        try:
            last_error: Exception | None = None
            for _ in range(retry_count + 1):
                try:
                    capture_result = await capture_screenshot(page, candidates, output_path, timeout_sec, wait_after_load)
                    final_url = capture_result.get("final_url", first_candidate)
                    actual_title = capture_result.get("actual_title", "")
                    final_path = finalize_screenshot_path(
                        current_path=output_path,
                        seq=asset.get("seq", ""),
                        fallback_title=asset.get("title", ""),
                        actual_title=actual_title,
                        final_url=final_url,
                    )
                    result = build_result_row(asset.get("seq", ""), target_value, final_url, RESULT_STATUS_SUCCESS, "", final_path)
                    if logger:
                        logger.info("Captured asset %s -> %s", asset.get("seq", ""), final_url)
                    return result
                except Exception as exc:
                    last_error = exc
            result = build_result_row(
                asset.get("seq", ""),
                target_value,
                first_candidate,
                RESULT_STATUS_FAILED,
                str(last_error) if last_error else "Unknown error",
                output_path,
            )
            if logger:
                logger.error("Failed asset %s -> %s: %s", asset.get("seq", ""), first_candidate, result["error"])
            return result
        finally:
            await page.close()


async def run_batch(assets: list[dict[str, str]], out_dir: Path, timeout_sec: int, wait_after_load: int, concurrency: int, retry_count: int, skip_existing: bool, headful: bool, logger: logging.Logger | None = None) -> list[dict[str, str]]:
    semaphore = asyncio.Semaphore(max(1, concurrency))
    out_dir.mkdir(parents=True, exist_ok=True)
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=not headful)
        context = await browser.new_context(viewport={"width": 1366, "height": 768})
        try:
            tasks = [
                run_one_asset(asset, context, semaphore, out_dir, timeout_sec, wait_after_load, retry_count, skip_existing, logger)
                for asset in assets
            ]
            results = await asyncio.gather(*tasks)
        finally:
            await context.close()
            await browser.close()
    return sorted(results, key=lambda item: (int(item["seq"]) if str(item.get("seq", "")).isdigit() else 0, item.get("seq", "")))


def run_job(config: dict[str, object], logger: logging.Logger | None = None) -> dict[str, Path | str]:
    runtime_logger = logger or setup_logger(Path(config["log_path"]))
    runtime_logger.info("Starting screenshot job")
    assets = load_assets_from_xlsx(Path(config["input_path"]), max_items=int(config["max_items"]))
    runtime_logger.info("Loaded %s assets", len(assets))
    results = asyncio.run(
        run_batch(
            assets=assets,
            out_dir=Path(config["output_dir"]),
            timeout_sec=int(config["timeout_sec"]),
            wait_after_load=int(config["wait_after_load"]),
            concurrency=int(config["concurrency"]),
            retry_count=int(config["retry_count"]),
            skip_existing=bool(config["skip_existing"]),
            headful=bool(config["headful"]),
            logger=runtime_logger,
        )
    )
    write_results_csv(results, Path(config["result_csv"]))
    summary = build_summary(results)
    summary_text = render_summary_text(summary)
    write_summary_text(Path(config["summary_txt"]), summary_text)
    runtime_logger.info("Completed screenshot job")
    return {
        "result_csv": Path(config["result_csv"]),
        "summary_txt": Path(config["summary_txt"]),
        "output_dir": Path(config["output_dir"]),
        "summary_text": summary_text,
    }


def main() -> None:
    args = build_parser().parse_args()
    config = build_run_config(
        input_path=Path(args.input),
        max_items=args.max_items,
        concurrency=args.concurrency,
        retry_count=args.retry_count,
        timeout_sec=args.timeout_sec,
        wait_after_load=args.wait_after_load,
        skip_existing=args.skip_existing,
        headful=args.headful,
    )
    print(f"Loaded configuration for {config['input_path']}")
    result = run_job(config)
    print(Path(result["summary_txt"]).read_text(encoding="utf-8"))
    print(f"Saved results to {result['result_csv']}")


if __name__ == "__main__":
    main()
