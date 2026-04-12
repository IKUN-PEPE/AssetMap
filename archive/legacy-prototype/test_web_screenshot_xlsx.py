from pathlib import Path

from openpyxl import Workbook

from asset_mapping_screenshot_xlsx.web_screenshot_xlsx import (
    build_candidate_urls,
    build_output_filename,
    build_output_path,
    build_parser,
    build_run_config,
    build_runtime_paths,
    build_summary,
    classify_failure_reason,
    finalize_screenshot_path,
    load_assets_from_xlsx,
    process_asset_row,
    render_summary_text,
    sanitize_filename,
    select_target_value,
    setup_logger,
    should_skip_existing,
    write_summary_text,
)


def test_build_runtime_paths_returns_log_and_output_paths(tmp_path: Path):
    paths = build_runtime_paths(base_dir=tmp_path)
    assert paths["log_path"] == tmp_path / "run.log"
    assert paths["output_dir"] == tmp_path / "screenshots"
    assert paths["result_csv"] == tmp_path / "results" / "results.csv"


def test_setup_logger_writes_to_run_log(tmp_path: Path):
    logger = setup_logger(tmp_path / "run.log")
    logger.info("hello log")
    for handler in logger.handlers:
        handler.flush()
    assert "hello log" in (tmp_path / "run.log").read_text(encoding="utf-8")


def test_build_run_config_merges_cli_like_values(tmp_path: Path):
    config = build_run_config(
        input_path=Path("input.xlsx"),
        base_dir=tmp_path,
        max_items=10,
        concurrency=3,
        retry_count=1,
        timeout_sec=20,
        wait_after_load=2,
        skip_existing=False,
        headful=True,
    )
    assert config["input_path"] == Path("input.xlsx")
    assert config["output_dir"] == tmp_path / "screenshots"
    assert config["skip_existing"] is False
    assert config["headful"] is True


def test_select_target_value_prefers_url_link_then_host():
    row = {"URL链接": "https://example.com/login", "域名/主机": "example.com"}
    assert select_target_value(row) == "https://example.com/login"


def test_select_target_value_falls_back_to_host_when_url_missing():
    row = {"URL链接": "", "域名/主机": "app.example.com"}
    assert select_target_value(row) == "app.example.com"


def test_select_target_value_returns_empty_when_both_missing():
    row = {"URL链接": "", "域名/主机": ""}
    assert select_target_value(row) == ""


def test_build_candidate_urls_keeps_full_url_as_single_candidate():
    assert build_candidate_urls("https://example.com/login") == ["https://example.com/login"]


def test_build_candidate_urls_generates_https_then_http_for_host():
    assert build_candidate_urls("portal.example.com") == [
        "https://portal.example.com",
        "http://portal.example.com",
    ]


def test_build_candidate_urls_ignores_blank_values():
    assert build_candidate_urls("") == []


def test_sanitize_filename_replaces_invalid_path_characters():
    assert sanitize_filename('A<B>:"C"/D\\E|F?*') == "A_B___C__D_E_F__"


def test_build_output_filename_contains_sequence_title_and_full_url():
    filename = build_output_filename(
        seq="001",
        page_title="登录页",
        final_url="https://example.com/login?a=1",
    )
    assert filename.startswith("001_登录页_https___example.com_login_a=1")
    assert filename.endswith(".png")


def test_classify_failure_reason_groups_timeout_and_other_errors():
    assert classify_failure_reason("Page.goto: Timeout 15000ms exceeded") == "timeout"
    assert classify_failure_reason("net::ERR_NAME_NOT_RESOLVED") == "dns_or_network"
    assert classify_failure_reason("some unexpected error") == "other"


def test_build_summary_aggregates_status_counts_and_rates():
    results = [
        {"status": "success", "error": ""},
        {"status": "failed", "error": "Page.goto: Timeout 15000ms exceeded"},
        {"status": "failed", "error": "net::ERR_NAME_NOT_RESOLVED"},
        {"status": "skipped", "error": ""},
    ]

    summary = build_summary(results)

    assert summary["total"] == 4
    assert summary["success"] == 1
    assert summary["failed"] == 2
    assert summary["skipped"] == 1
    assert summary["success_rate"] == "25.00%"
    assert summary["failure_reasons"] == {"timeout": 1, "dns_or_network": 1}


def test_render_summary_text_contains_counts_and_failure_reasons():
    summary = {
        "total": 4,
        "success": 1,
        "failed": 2,
        "skipped": 1,
        "success_rate": "25.00%",
        "failed_rate": "50.00%",
        "skipped_rate": "25.00%",
        "failure_reasons": {"timeout": 1, "dns_or_network": 1},
    }

    text = render_summary_text(summary)

    assert "Total: 4" in text
    assert "Success: 1 (25.00%)" in text
    assert "Failed: 2 (50.00%)" in text
    assert "Skipped: 1 (25.00%)" in text
    assert "- timeout: 1" in text
    assert "- dns_or_network: 1" in text


def test_render_summary_text_sorts_failure_reasons():
    summary = {
        "total": 3,
        "success": 0,
        "failed": 3,
        "skipped": 0,
        "success_rate": "0.00%",
        "failed_rate": "100.00%",
        "skipped_rate": "0.00%",
        "failure_reasons": {"timeout": 1, "dns_or_network": 1, "other": 1},
    }

    text = render_summary_text(summary)
    assert text.index("- dns_or_network: 1") < text.index("- other: 1") < text.index("- timeout: 1")


def test_write_summary_text_persists_summary_file(tmp_path: Path):
    summary_path = tmp_path / "summary.txt"
    write_summary_text(summary_path, "Total: 3\nSuccess: 2 (66.67%)")
    assert summary_path.read_text(encoding="utf-8") == "Total: 3\nSuccess: 2 (66.67%)"


def test_finalize_screenshot_path_renames_with_actual_title_and_overwrites_target(tmp_path: Path):
    initial_path = tmp_path / "1_旧标题_http___example.com.png"
    initial_path.write_bytes(b"new-image")
    existing_target = tmp_path / "1_真实标题_http___example.com.png"
    existing_target.write_bytes(b"old-image")

    final_path = finalize_screenshot_path(
        current_path=initial_path,
        seq="1",
        fallback_title="旧标题",
        actual_title="真实标题",
        final_url="http://example.com",
    )

    assert final_path == existing_target
    assert final_path.read_bytes() == b"new-image"
    assert not initial_path.exists()


def test_load_assets_from_xlsx_reads_expected_headers(tmp_path: Path):
    workbook_path = tmp_path / "assets.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["序号", "域名/主机", "网页标题", "URL链接"])
    sheet.append([1, "example.com", "首页", "https://example.com"])
    workbook.save(workbook_path)

    assets = load_assets_from_xlsx(workbook_path, max_items=10)

    assert assets == [
        {
            "seq": "1",
            "host": "example.com",
            "title": "首页",
            "url": "https://example.com",
        }
    ]


def test_load_assets_from_xlsx_skips_leading_blank_rows_before_header(tmp_path: Path):
    workbook_path = tmp_path / "assets_with_blank_rows.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([None, None, None, None])
    sheet.append(["说明", None, None, None])
    sheet.append(["序号", "域名/主机", "网页标题", "URL链接"])
    sheet.append([2, "blank.example.com", "空行后", "https://blank.example.com"])
    workbook.save(workbook_path)

    assets = load_assets_from_xlsx(workbook_path, max_items=10)

    assert assets == [
        {
            "seq": "2",
            "host": "blank.example.com",
            "title": "空行后",
            "url": "https://blank.example.com",
        }
    ]


def test_load_assets_from_xlsx_keeps_all_rows_when_max_items_zero(tmp_path: Path):
    workbook_path = tmp_path / "all_rows.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["序号", "域名/主机", "网页标题", "URL链接"])
    sheet.append([1, "one.example.com", "One", "https://one.example.com"])
    sheet.append([2, "two.example.com", "Two", "https://two.example.com"])
    workbook.save(workbook_path)

    assets = load_assets_from_xlsx(workbook_path, max_items=0)

    assert [asset["seq"] for asset in assets] == ["1", "2"]


def test_should_skip_existing_returns_true_when_file_exists(tmp_path: Path):
    output_path = build_output_path(tmp_path, "001", "首页", "https://example.com")
    output_path.write_bytes(b"png")
    assert should_skip_existing(output_path) is True


def test_process_asset_row_uses_actual_title_in_success_path(tmp_path):
    asset = {"seq": "1", "host": "example.com", "title": "旧标题", "url": "https://example.com"}

    def fake_capture(candidates, output_path):
        assert candidates == ["https://example.com"]
        output_path.write_bytes(b"png")
        return {
            "final_url": "https://example.com",
            "actual_title": "真实标题",
        }

    result = process_asset_row(asset=asset, out_dir=tmp_path, capture_func=fake_capture, skip_existing=False)

    assert result["status"] == "success"
    assert result["final_url"] == "https://example.com"
    assert result["screenshot_path"].endswith("1_真实标题_https___example.com.png")


def test_process_asset_row_keeps_excel_title_when_actual_title_is_blank(tmp_path: Path):
    asset = {"seq": "2", "host": "example.com", "title": "Excel标题", "url": "https://example.com"}

    def fake_capture(candidates, output_path):
        output_path.write_bytes(b"png")
        return {
            "final_url": "https://example.com",
            "actual_title": "",
        }

    result = process_asset_row(asset=asset, out_dir=tmp_path, capture_func=fake_capture, skip_existing=False)

    assert result["status"] == "success"
    assert result["screenshot_path"].endswith("2_Excel标题_https___example.com.png")


def test_build_parser_sets_batch_defaults():
    args = build_parser().parse_args([])
    assert args.concurrency == 5
    assert args.retry_count == 1
    assert args.skip_existing is True


def test_build_parser_uses_project_local_output_defaults():
    args = build_parser().parse_args([])
    normalized_output_dir = str(args.output_dir).replace('\\', '/')
    normalized_result_csv = str(args.result_csv).replace('\\', '/')
    assert normalized_output_dir.endswith('asset_mapping_screenshot_xlsx/screenshots')
    assert normalized_result_csv.endswith('asset_mapping_screenshot_xlsx/results/results.csv')
